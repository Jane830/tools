import json
import openpyxl
import db_utils


# ivr, the first sheet
def write_content_ivr(timestamp, filename):
    content_list = db_utils.connect_newfs_bt_autocall(timestamp, '%ivr')
    j = 2
    for content in content_list:
        j = j + 1
        json_data = json.loads(content)
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.worksheets[0]
        for i in range(1, 26, 1):
            data = sheet.cell(i, 1).value
            if data in json_data:
                data_value = json_data[data]
                if data_value is not None:
                    sheet.cell(i, j).value = data_value
        workbook.save(filename)


# bot, the second sheet
def write_content_bot(timestamp, filename):
    content_list = db_utils.connect_newfs_bt_autocall(timestamp, '%print')
    j = 2
    for content in content_list:
        j = j + 1
        json_data = json.loads(content)
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.worksheets[1]
        for i in range(1, 29, 1):
            data = sheet.cell(i, 1).value
            if data in json_data:
                data_value = json_data[data]
                if data_value is not None:
                    sheet.cell(i, j).value = data_value
        workbook.save(filename)


# agent, the third sheet
def write_content_agent(timestamp, filename):
    content_list = db_utils.connect_newfs_bt_autocall(timestamp, '%agent')
    j = 2
    for content in content_list:
        j = j + 1
        json_data = json.loads(content)
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.worksheets[2]
        for i in range(1, 28, 1):
            data = sheet.cell(i, 1).value
            if data in json_data:
                data_value = json_data[data]
                if data_value is not None:
                    sheet.cell(i, j).value = data_value
        workbook.save(filename)


def write_content_manual(timestamp, filename):
    content_list = db_utils.connect_newfs_bt_autocall(timestamp, '%call')
    write_content(filename, content_list)
    write_content2(filename, content_list)


def write_content(filename, content_list):
    j = 2
    for content in content_list:
        j = j + 1
        json_data = json.loads(content)
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.worksheets[0]  # .get_sheet_by_name('Sheet1')
        for i in range(1, 14, 1):
            data = sheet.cell(i, 1).value
            if data in json_data:
                data_value = json_data[data]
                if data_value is not None:
                    sheet.cell(i, j).value = data_value
        workbook.save(filename)


def write_content2(filename, content_list):
    j = 2
    for content in content_list:
        j = j + 1
        json_data = json.loads(content)
        callid = json_data['callid']
        devices_list = json_data['devices']
        for m in devices_list:
            m = json.dumps(m)
            new_json_data = json.loads(m)
            workbook = openpyxl.load_workbook(filename)
            sheet = workbook.worksheets[1]
            for i in range(1, 47, 1):
                sheet.cell(1, j).value = callid
                data = sheet.cell(i, 1).value
                if data in new_json_data:
                    data_value = new_json_data[data]
                    if data_value is not None:
                        sheet.cell(i, j).value = data_value
            j = j + 1
            workbook.save(filename)


if __name__ == '__main__':
    '''    
    预测外呼 转ivr & 转bot & 转agent
    ivr 在第一个sheet中；bot在第二个sheet中；agent在第三个sheet中;
    测试的时候可同时执行一下函数，同时写入demo_auto.xlsx
    
    '''
    # write_content_ivr(timestamp=1620357073, filename='demo_auto.xlsx')
    # write_content_bot(timestamp=1620357073, filename='demo_auto.xlsx')
    # write_content_agent(timestamp=1620357073, filename='demo_auto.xlsx')

    '''
    手动外呼 & 呼入
    由于手动含多侧话单，第一个sheet为最外层数据；第二个sheet为内层device数据
    '''
    write_content_manual(timestamp=1620357073, filename='demo_manual.xlsx')

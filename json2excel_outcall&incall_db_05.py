import json
import openpyxl
from collections import OrderedDict
from utils_push import db_utils


def to_order_dict(origin_dict, custom_list):
    for key1, value1 in origin_dict.items():  # 当字典的key值为时间的话
        if key1 in custom_list:  # 将字典中的值保存到字典中，筛选需要的字典
            origin_dict[key1] = value1
    order_dict = OrderedDict()
    for key in custom_list:  # 获取orderdict，自定义顺序的字典
        if key in origin_dict.keys():
            order_dict[key] = origin_dict.get(key)
    return order_dict


def write_order_dict(sheet, order_dict, y):
    x = 0  # 循环字典，打印值
    for key, value in order_dict.items():
        sheet.write(x, y, key)
        sheet.write(x, y + 1, value)
        x = x + 1


def get_time(book, json_data, sheet_name, field_list):
    sheet_time = book.add_sheet(sheet_name, cell_overwrite_ok=True)  # 添加一个sheet页
    order_dict_01 = to_order_dict(json_data, field_list)
    write_order_dict(sheet_time, order_dict_01, 0)
    for key, value in json_data.items():
        if isinstance(value, list):
            y = 2
            for index, inner_dict in enumerate(value):  # 获取list中数据
                order_dict = to_order_dict(inner_dict, field_list)
                write_order_dict(sheet_time, order_dict, y)
                y = y + 2






def write_content(timestamp):
    content_list = db_utils.connect_newfs_bt(timestamp)
    j = 2
    for content in content_list:
        j = j + 1
        json_data = json.loads(content)
        workbook = openpyxl.load_workbook('demo01.xlsx')
        sheet = workbook.active
        sheet = workbook.worksheets[0]       #.get_sheet_by_name('Sheet1')
        for i in range(1,14,1):
            data=sheet.cell(i, 1).value
            if data in json_data:
                data_value=json_data[data]
                if data_value is not None:
                    sheet.cell(i,j).value=data_value
        workbook.save('demo01.xlsx')


def write_content2(timestamp):
    content_list = db_utils.connect_newfs_bt(timestamp)
    j = 2
    for content in content_list:
        j=j+1
        json_data = json.loads(content)
        callid=json_data['callid']
        devices_list = json_data['devices']
        for m in devices_list:
            m=json.dumps(m)
            new_json_data = json.loads(m)
            workbook = openpyxl.load_workbook('demo01.xlsx')
            sheet = workbook.active
            sheet = workbook.worksheets[1]
            for i in range(1,47,1):
                sheet.cell(1, j).value = callid
                data=sheet.cell(i, 1).value
                if data in new_json_data:
                    data_value = new_json_data[data]
                    if data_value is not None:
                        sheet.cell(i, j).value = data_value
            j=j+1

            workbook.save('demo01.xlsx')




if __name__ == '__main__':
    write_content(timestamp=1620464550)
    write_content2(timestamp=1620464550)

